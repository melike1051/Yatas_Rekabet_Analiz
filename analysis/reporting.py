from __future__ import annotations

import mimetypes
import os
import smtplib
import json
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import pandas as pd
from sqlalchemy import text

from analysis.executive_summary import generate_executive_summary
from db.session import engine
from scraper.utils.dining_room import (
    ITEM_TYPE_ORDER,
    build_match_key as build_dining_room_match_key,
    classify_product,
    clean_product_name,
    derive_team_name as derive_dining_room_team_name,
    infer_item_type,
    is_team_row as is_dining_room_team_row,
)
from scraper.utils.logging_config import get_logger
from scraper.utils.normalizers import dump_json


logger = get_logger("analysis.reporting")
REPORTS_DIR = Path("analysis/data/reports")
REPORT_METADATA_PATH = REPORTS_DIR / "latest_report.json"
MATCH_OVERRIDES_PATH = Path("analysis/data/report_match_overrides.json")
BRAND_ORDER = ["istikbal", "bellona", "dogtas"]


def _excel_column_letter(index: int) -> str:
    result = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def _price_pair(list_candidate: Any, discount_candidate: Any) -> tuple[Any, Any]:
    values = [value for value in (list_candidate, discount_candidate) if pd.notna(value)]
    if not values:
        return None, None
    if len(values) == 1:
        return values[0], None
    high_value = max(values)
    low_value = min(values)
    if high_value == low_value:
        return high_value, None
    return high_value, low_value


def _sum_series(values: pd.Series) -> Any:
    non_null_values = values.dropna()
    if non_null_values.empty:
        return None
    return non_null_values.sum()


def _style_assets() -> dict[str, Any]:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "Alignment": Alignment,
        "Font": Font,
        "HEADER_FILL": PatternFill(fill_type="solid", fgColor="9BBB59"),
        "SECTION_FILL": PatternFill(fill_type="solid", fgColor="D9EAD3"),
        "SUBHEADER_FILL": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "ITEM_FILL": PatternFill(fill_type="solid", fgColor="F9CB9C"),
        "TEAM_FILL": PatternFill(fill_type="solid", fgColor="6AA84F"),
        "PRICE_FILL": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "THIN_BORDER": Border(
            left=Side(style="thin", color="D0D7DE"),
            right=Side(style="thin", color="D0D7DE"),
            top=Side(style="thin", color="D0D7DE"),
            bottom=Side(style="thin", color="D0D7DE"),
        ),
    }


def parse_email_recipients(raw_value: str | None) -> list[str]:
    if not raw_value:
        return []
    normalized = raw_value.replace(";", ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def flatten_catalog_diff_rows(summary: dict[str, Any] | None) -> list[dict[str, Any]]:
    brands = ((summary or {}).get("catalog_diff_summary") or {}).get("brands") or {}
    rows: list[dict[str, Any]] = []
    for brand, payload in brands.items():
        summary_block = payload.get("summary", {})
        rows.append(
            {
                "Marka": brand,
                "Durum": payload.get("status"),
                "Onceki Katalog": summary_block.get("previous_count", 0),
                "Guncel Katalog": summary_block.get("current_count", 0),
                "Yeni Urun": summary_block.get("new_count", 0),
                "Kalkan Urun": summary_block.get("removed_count", 0),
                "Sabit Kalan": summary_block.get("unchanged_count", 0),
            }
        )
    return rows


def build_management_summary(summary: dict[str, Any]) -> list[str]:
    overview = summary.get("overview", {})
    price_summary = summary.get("price_summary", {})
    promotion_summary = summary.get("promotion_summary", {})
    ai_insights = summary.get("ai_insights", {})
    diff_rows = flatten_catalog_diff_rows(summary)
    total_new_products = sum(int(row.get("Yeni Urun", 0) or 0) for row in diff_rows)
    total_removed_products = sum(int(row.get("Kalkan Urun", 0) or 0) for row in diff_rows)
    top_discount_brand = price_summary.get("top_discount_brand") or "Veri yok"
    promotion_brands = promotion_summary.get("brands", [])
    leading_promotion = promotion_brands[0] if promotion_brands else {}
    promotion_line = (
        f"Kampanya tarafinda {leading_promotion.get('competitor_name', 'veri yok')} "
        f"markasi one cikiyor; sepette {leading_promotion.get('basket_discount_count', 0)} "
        f"kampanya ve toplam {leading_promotion.get('promotion_count', 0)} mesaj izlendi."
        if promotion_brands
        else "Kampanya tarafinda anlamli mesaj siniflandirmasi henuz olusmadi."
    )

    lines = [
        (
            f"Izlenen {overview.get('competitor_count', 0)} rakipte toplam "
            f"{overview.get('product_count', 0)} urun aktif olarak takip ediliyor."
        ),
        (
            f"Son 7 gunde {overview.get('weekly_promotion_count', 0)} kampanya ve "
            f"{overview.get('out_of_stock_count', 0)} stok riski tespit edildi."
        ),
        promotion_line,
        (
            f"Fiyat aksiyonunda en agresif marka {top_discount_brand}; "
            f"{price_summary.get('price_decreased_count', 0)} urunde fiyat dususu izlendi."
        ),
        f"Haftalik katalog hareketinde {total_new_products} yeni, {total_removed_products} pasif urun tespit edildi.",
    ]
    if ai_insights.get("strategic_summary"):
        lines.append(str(ai_insights["strategic_summary"]))
    return lines


def infer_product_type(product_name: str | None) -> str:
    return infer_item_type(product_name)


def derive_collection_name(product_name: str | None) -> str:
    return derive_dining_room_team_name(product_name)


def build_match_key(product_name: str | None, product_type: str | None) -> str:
    return build_dining_room_match_key(product_name, product_type, derive_collection_name(product_name))


def _normalize_override_text(value: str | None) -> str:
    return str(value or "").strip().casefold()


def load_match_overrides() -> dict[str, list[dict[str, Any]] | dict[str, dict[str, Any]]]:
    if not MATCH_OVERRIDES_PATH.exists():
        return {"sku": {}, "pattern": []}
    try:
        payload = json.loads(MATCH_OVERRIDES_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"sku": {}, "pattern": []}

    if not isinstance(payload, list):
        return {"sku": {}, "pattern": []}
    overrides: dict[str, dict[str, Any]] = {}
    pattern_overrides: list[dict[str, Any]] = []
    if not payload:
        return {"sku": overrides, "pattern": pattern_overrides}
    for row in payload:
        if not isinstance(row, dict):
            continue
        sku = row.get("competitor_sku")
        if sku:
            overrides[str(sku)] = row
            continue
        if row.get("brand") or row.get("name_contains") or row.get("team_name_contains"):
            pattern_overrides.append(row)
    return {"sku": overrides, "pattern": pattern_overrides}


def _resolve_override(
    override_payload: dict[str, list[dict[str, Any]] | dict[str, dict[str, Any]]],
    competitor_sku: str,
    competitor_name: str,
    product_name: str,
    raw_team_name: str | None,
) -> dict[str, Any]:
    sku_override = (override_payload.get("sku") or {}).get(str(competitor_sku))
    if isinstance(sku_override, dict):
        return sku_override

    normalized_brand = _normalize_override_text(competitor_name)
    normalized_product = _normalize_override_text(product_name)
    normalized_team = _normalize_override_text(raw_team_name)

    for row in override_payload.get("pattern") or []:
        if not isinstance(row, dict):
            continue
        brand = row.get("brand")
        if brand and _normalize_override_text(brand) != normalized_brand:
            continue
        name_contains = row.get("name_contains")
        if name_contains and _normalize_override_text(name_contains) not in normalized_product:
            continue
        team_contains = row.get("team_name_contains")
        if team_contains and _normalize_override_text(team_contains) not in normalized_team:
            continue
        return row
    return {}


def _choose_team_name(raw_team_name: str | None, classified_team_name: str) -> str:
    if not raw_team_name:
        return classified_team_name
    normalized_raw = _normalize_override_text(raw_team_name)
    normalized_classified = _normalize_override_text(classified_team_name)
    if not normalized_classified:
        return str(raw_team_name)
    if normalized_raw.startswith(normalized_classified) and len(normalized_classified) < len(normalized_raw):
        return classified_team_name
    if any(char.isdigit() for char in str(raw_team_name)):
        return classified_team_name
    return str(raw_team_name)


def _load_product_specs_report_frame(limit: int = 250) -> pd.DataFrame:
    query = text(
        """
        SELECT
            c.name AS competitor_name,
            p.product_name,
            p.competitor_sku,
            p.current_price,
            ps.material_type,
            ps.tabletop_thickness_mm,
            ps.width_cm,
            ps.depth_cm,
            ps.height_cm,
            ps.skeleton_type,
            ps.color,
            ps.parsed_by,
            ps.confidence_score
        FROM product_specs ps
        JOIN products p ON p.id = ps.product_id
        JOIN competitors c ON c.id = p.competitor_id
        ORDER BY ps.updated_at DESC
        LIMIT :limit
        """
    )
    with engine.connect() as connection:
        return pd.read_sql_query(query, connection, params={"limit": limit})


def _load_manual_report_base_frame() -> pd.DataFrame:
    query = text(
        """
        WITH latest_prices AS (
            SELECT
                ph.product_id,
                ph.price,
                ph.original_price,
                ph.discount_rate,
                ph.promotion_label,
                ph.captured_at,
                ROW_NUMBER() OVER (
                    PARTITION BY ph.product_id
                    ORDER BY ph.captured_at DESC
                ) AS row_num
            FROM price_history ph
        )
        SELECT
            c.name AS competitor_name,
            p.product_name,
            p.competitor_sku,
            p.product_url,
            p.category_name,
            p.current_price,
            lp.price AS latest_price,
            lp.original_price,
            lp.discount_rate,
            lp.promotion_label,
            lp.captured_at AS latest_captured_at,
            p.raw_attributes,
            ps.material_type,
            ps.color,
            ps.parsed_by,
            ps.confidence_score
        FROM products p
        JOIN competitors c ON c.id = p.competitor_id
        LEFT JOIN latest_prices lp ON lp.product_id = p.id AND lp.row_num = 1
        LEFT JOIN product_specs ps ON ps.product_id = p.id
        ORDER BY c.name, p.product_name
        """
    )
    with engine.connect() as connection:
        frame = pd.read_sql_query(query, connection)

    if frame.empty:
        return frame

    overrides = load_match_overrides()
    frame["raw_attributes"] = frame["raw_attributes"].apply(lambda value: value or {})

    def enrich_row(row: pd.Series) -> pd.Series:
        raw = dict(row["raw_attributes"] or {})
        cleaned_name = clean_product_name(row["product_name"])
        classification = classify_product(cleaned_name)
        override = _resolve_override(
            overrides,
            str(row["competitor_sku"]),
            str(row["competitor_name"]),
            cleaned_name,
            raw.get("team_name"),
        )
        raw_item_type = raw.get("item_type")
        item_type = override.get("item_type") or (
            raw_item_type if raw_item_type in ITEM_TYPE_ORDER else classification["item_type"]
        )
        raw_team_name = raw.get("team_name")
        canonical_team_name = _choose_team_name(raw_team_name, str(classification["team_name"]))
        team_name = override.get("team_name") or (
            canonical_team_name
            if raw_item_type in ITEM_TYPE_ORDER and canonical_team_name
            else str(classification["team_name"])
        )
        is_team = bool(
            override.get("is_team_row")
            if "is_team_row" in override
            else raw.get("is_team_row") or classification["is_team_row"]
        )
        team_size_variant = override.get("team_size_variant") or raw.get("team_size_variant") or classification["team_size_variant"]
        display_order = int(override.get("display_order") or raw.get("display_order") or ITEM_TYPE_ORDER.get(str(item_type), 999))
        match_group = override.get("match_group") or raw.get("match_group") or build_dining_room_match_key(
            cleaned_name,
            str(item_type),
            str(team_name),
        )
        row["Temiz Urun Adi"] = cleaned_name
        row["Takim"] = team_name
        row["Urun Cesidi"] = item_type
        row["Takim Satiri"] = is_team
        row["Takim Varyanti"] = team_size_variant
        row["Display Order"] = display_order
        row["Match Group"] = match_group
        row["Team Display Name"] = team_size_variant or cleaned_name
        return row

    frame = frame.apply(enrich_row, axis=1)
    price_pairs = frame.apply(
        lambda row: _price_pair(
            row["latest_price"] if pd.notna(row["latest_price"]) else row["current_price"],
            row["original_price"],
        ),
        axis=1,
        result_type="expand",
    )
    frame["Liste Fiyat"] = price_pairs[0]
    frame["Ind. PRK Fiyat"] = price_pairs[1]
    frame["Son Fiyat"] = frame["latest_price"].fillna(frame["current_price"])
    frame["Marka"] = frame["competitor_name"].str.upper()
    frame["Son Fiyat Tarihi"] = pd.to_datetime(frame["latest_captured_at"]).dt.strftime("%d.%m.%Y")
    frame["Indirim %"] = frame["discount_rate"]
    return frame.loc[
        frame["Urun Cesidi"].isin({"Yemek Odasi", "Konsol", "Sabit Masa", "Acilir Masa", "Sandalye", "Bench / Puf", "Vitrin"})
    ].copy()


def _load_monthly_snapshot_frame(limit_months: int = 3) -> tuple[pd.DataFrame, list[str]]:
    month_query = text(
        """
        WITH monthly_dates AS (
            SELECT
                date_trunc('month', captured_at) AS month_bucket,
                MAX(captured_at::date) AS snapshot_date
            FROM price_history
            GROUP BY 1
            ORDER BY month_bucket DESC
            LIMIT :limit_months
        )
        SELECT snapshot_date
        FROM monthly_dates
        ORDER BY snapshot_date ASC
        """
    )
    with engine.connect() as connection:
        snapshot_dates = [row.snapshot_date for row in connection.execute(month_query, {"limit_months": limit_months}).all()]
        if not snapshot_dates:
            return pd.DataFrame(), []
        snapshot_query = text(
            """
            WITH ranked_monthly_prices AS (
                SELECT
                    ph.product_id,
                    ph.captured_at::date AS snapshot_date,
                    ph.price,
                    ph.original_price,
                    ROW_NUMBER() OVER (
                        PARTITION BY ph.product_id, ph.captured_at::date
                        ORDER BY ph.captured_at DESC
                    ) AS row_num
                FROM price_history ph
                WHERE ph.captured_at::date = ANY(:snapshot_dates)
            )
            SELECT
                c.name AS competitor_name,
                p.product_name,
                p.competitor_sku,
                rmp.snapshot_date,
                rmp.price,
                rmp.original_price
            FROM ranked_monthly_prices rmp
            JOIN products p ON p.id = rmp.product_id
            JOIN competitors c ON c.id = p.competitor_id
            WHERE rmp.row_num = 1
            ORDER BY c.name, p.product_name, rmp.snapshot_date
            """
        )
        frame = pd.read_sql_query(snapshot_query, connection, params={"snapshot_dates": snapshot_dates})

    return frame, [pd.to_datetime(date).strftime("%d.%m.%Y") for date in snapshot_dates]


def _build_panel_frame(base_frame: pd.DataFrame, snapshot_frame: pd.DataFrame, snapshot_columns: list[str]) -> pd.DataFrame:
    if base_frame.empty:
        return pd.DataFrame()

    panel = base_frame.copy()
    panel = panel[
        [
            "Marka",
            "competitor_sku",
            "Takim",
            "Urun Cesidi",
            "Team Display Name",
            "Display Order",
            "Takim Satiri",
            "product_url",
            "Liste Fiyat",
            "Ind. PRK Fiyat",
        ]
    ].rename(
        columns={
            "competitor_sku": "Malzeme Kodu",
            "Team Display Name": "Urun Adi",
            "product_url": "Urun Linki",
            "Liste Fiyat": "Guncel Liste fiyat",
            "Ind. PRK Fiyat": "Guncel IND. PRK FIYAT",
        }
    )

    if not snapshot_frame.empty:
        snapshot_frame = snapshot_frame.copy()
        snapshot_frame["snapshot_date"] = pd.to_datetime(snapshot_frame["snapshot_date"]).dt.strftime("%d.%m.%Y")
        snapshot_price_pairs = snapshot_frame.apply(
            lambda row: _price_pair(row["price"], row["original_price"]),
            axis=1,
            result_type="expand",
        )
        snapshot_frame["Liste fiyat"] = snapshot_price_pairs[0]
        snapshot_frame["IND. PRK FIYAT"] = snapshot_price_pairs[1]

        merged = panel.merge(
            snapshot_frame[["competitor_sku", "snapshot_date", "Liste fiyat", "IND. PRK FIYAT"]],
            left_on="Malzeme Kodu",
            right_on="competitor_sku",
            how="left",
        )
        pivot = (
            merged.pivot_table(
                index=["Malzeme Kodu"],
                columns="snapshot_date",
                values=["Liste fiyat", "IND. PRK FIYAT"],
                aggfunc="last",
            )
            if not merged.empty
            else pd.DataFrame()
        )
        if not pivot.empty:
            pivot.columns = [f"{date}|{metric}" for metric, date in pivot.columns]
            pivot = pivot.reset_index()
            panel = panel.merge(pivot, on="Malzeme Kodu", how="left")

    for date_label in snapshot_columns:
        for metric in ("Liste fiyat", "IND. PRK FIYAT"):
            column_name = f"{date_label}|{metric}"
            if column_name not in panel.columns:
                panel[column_name] = None

    if snapshot_columns:
        latest_snapshot = snapshot_columns[-1]
        latest_list_column = f"{latest_snapshot}|Liste fiyat"
        latest_discount_column = f"{latest_snapshot}|IND. PRK FIYAT"
        panel[latest_list_column] = panel[latest_list_column].where(
            panel[latest_list_column].notna(),
            panel["Guncel Liste fiyat"],
        )
        panel[latest_discount_column] = panel[latest_discount_column].where(
            panel[latest_discount_column].notna(),
            panel["Guncel IND. PRK FIYAT"],
        )

    ordered_dynamic_columns = [
        f"{date_label}|{metric}"
        for date_label in snapshot_columns
        for metric in ("Liste fiyat", "IND. PRK FIYAT")
    ]
    panel = panel.sort_values(
        ["Marka", "Takim", "Display Order", "Takim Satiri", "Urun Adi"],
        ascending=[True, True, True, True, True],
    ).reset_index(drop=True)
    panel = _synthesize_team_total_rows(panel, ordered_dynamic_columns)
    panel = _synthesize_team_summary_rows(panel, ordered_dynamic_columns)
    panel = panel.drop(columns=["Guncel Liste fiyat", "Guncel IND. PRK FIYAT"])
    return panel[
        ["Marka", "Malzeme Kodu", "Takim", "Urun Cesidi", "Urun Adi", "Display Order", "Takim Satiri", "Urun Linki", *ordered_dynamic_columns]
    ]


def _synthesize_team_total_rows(panel: pd.DataFrame, dynamic_columns: list[str]) -> pd.DataFrame:
    if panel.empty:
        return panel

    component_rows = panel.loc[~panel["Takim Satiri"]].copy()
    if component_rows.empty:
        return panel

    synthetic_rows: list[dict[str, Any]] = []
    for (brand, team_name), group in component_rows.groupby(["Marka", "Takim"], dropna=False):
        if not team_name:
            continue
        row = {column: None for column in panel.columns}
        row["Marka"] = brand
        row["Malzeme Kodu"] = f"{brand}::{team_name}::team-total"
        row["Takim"] = team_name
        row["Urun Cesidi"] = "Yemek Odasi"
        row["Urun Adi"] = "Takim Toplami"
        row["Display Order"] = ITEM_TYPE_ORDER["Yemek Odasi"]
        row["Takim Satiri"] = True
        row["Urun Linki"] = group["Urun Linki"].dropna().iloc[0] if group["Urun Linki"].notna().any() else None
        row["Guncel Liste fiyat"] = _sum_series(group["Guncel Liste fiyat"])
        row["Guncel IND. PRK FIYAT"] = _sum_series(group["Guncel IND. PRK FIYAT"])
        for column in dynamic_columns:
            row[column] = _sum_series(group[column]) if column in group.columns else None
        synthetic_rows.append(row)

    if not synthetic_rows:
        return panel

    synthetic_frame = pd.DataFrame(synthetic_rows).reindex(columns=panel.columns)
    combined = pd.concat([panel, synthetic_frame], ignore_index=True, sort=False)
    return combined.sort_values(
        ["Marka", "Takim", "Display Order", "Takim Satiri", "Urun Adi"],
        ascending=[True, True, True, True, True],
    ).reset_index(drop=True)


def _synthesize_team_summary_rows(panel: pd.DataFrame, dynamic_columns: list[str]) -> pd.DataFrame:
    if panel.empty:
        return panel

    team_rows = panel.loc[
        panel["Takim Satiri"]
        & ~panel["Urun Adi"].isin(["Takim Toplami", "Takim (1) Min", "Takim (2) Max"])
    ].copy()
    if team_rows.empty:
        return panel

    non_team_rows = panel.loc[~panel["Takim Satiri"]].copy()
    synthetic_rows: list[dict[str, Any]] = []

    for (brand, team_name), group in team_rows.groupby(["Marka", "Takim"], dropna=False):
        ranking_metric = group["Guncel IND. PRK FIYAT"].fillna(group["Guncel Liste fiyat"])
        ranked = group.assign(_ranking_metric=ranking_metric).sort_values(
            ["_ranking_metric", "Urun Adi"],
            ascending=[True, True],
        )
        if ranked.empty:
            continue

        def _build_summary_row(source_row: pd.Series, label: str, suffix: str) -> dict[str, Any]:
            row = source_row.to_dict()
            row["Malzeme Kodu"] = f"{source_row['Malzeme Kodu']}::{suffix}"
            row["Urun Adi"] = label
            row["Urun Cesidi"] = "Yemek Odasi"
            row["Takim Satiri"] = True
            row["Display Order"] = ITEM_TYPE_ORDER["Yemek Odasi"]
            return row

        min_row = ranked.iloc[0]
        synthetic_rows.append(_build_summary_row(min_row, "Takim (1) Min", "min"))

        max_row = ranked.iloc[-1]
        if len(ranked) > 1 and max_row["Malzeme Kodu"] != min_row["Malzeme Kodu"]:
            synthetic_rows.append(_build_summary_row(max_row, "Takim (2) Max", "max"))

    if not synthetic_rows:
        return panel

    synthetic_frame = pd.DataFrame(synthetic_rows)
    synthetic_frame = synthetic_frame.reindex(columns=panel.columns)
    non_team_rows = non_team_rows.reindex(columns=panel.columns)
    combined = pd.concat([non_team_rows, synthetic_frame], ignore_index=True, sort=False)
    combined = combined.sort_values(
        ["Marka", "Takim", "Display Order", "Takim Satiri", "Urun Adi"],
        ascending=[True, True, True, True, True],
    ).reset_index(drop=True)
    return combined


def _build_price_band_frame(base_frame: pd.DataFrame) -> pd.DataFrame:
    if base_frame.empty:
        return pd.DataFrame()

    price_frame = base_frame.loc[base_frame["Takim Satiri"]].copy()
    if price_frame.empty:
        price_frame = base_frame.copy()
    price_frame["Band Fiyati"] = price_frame["Ind. PRK Fiyat"].fillna(price_frame["Liste Fiyat"]).fillna(price_frame["Son Fiyat"])
    bins = [0, 20000, 30000, 40000, 50000, 60000, 80000, float("inf")]
    labels = ["0-20.000", "20.000-30.000", "30.000-40.000", "40.000-50.000", "50.000-60.000", "60.000-80.000", "80.000+"]
    price_frame["Fiyat Araligi"] = pd.cut(price_frame["Band Fiyati"], bins=bins, labels=labels, right=False)
    pivot = (
        price_frame.pivot_table(
            index="Fiyat Araligi",
            columns="Marka",
            values="competitor_sku",
            aggfunc="count",
            fill_value=0,
        )
        .reset_index()
        .rename_axis(None, axis=1)
    )
    if pivot.empty:
        return pivot
    for brand in [brand.upper() for brand in BRAND_ORDER]:
        if brand not in pivot.columns:
            pivot[brand] = 0
    brand_columns = [brand.upper() for brand in BRAND_ORDER]
    pivot["TOPLAM"] = pivot[brand_columns].sum(axis=1)
    return pivot[["Fiyat Araligi", *brand_columns, "TOPLAM"]]


def build_manual_report_context(limit_months: int = 3) -> dict[str, pd.DataFrame | list[str]]:
    base_frame = _load_manual_report_base_frame()
    snapshot_frame, snapshot_columns = _load_monthly_snapshot_frame(limit_months=limit_months)
    return {
        "base": base_frame,
        "panel": _build_panel_frame(base_frame, snapshot_frame, snapshot_columns),
        "price_bands": _build_price_band_frame(base_frame),
        "snapshot_columns": snapshot_columns,
    }


def _load_price_changes_report_frame(summary: dict[str, Any]) -> pd.DataFrame:
    frame = pd.DataFrame(summary.get("latest_price_changes", []))
    if frame.empty:
        return frame
    return frame.rename(
        columns={
            "competitor_name": "Marka",
            "product_name": "Urun",
            "competitor_sku": "SKU",
            "current_price": "Guncel Fiyat",
            "previous_price": "Onceki Fiyat",
            "price_change": "Degisim",
            "captured_at": "Yakalandi",
        }
    )


def _load_promotion_report_frame(summary: dict[str, Any]) -> pd.DataFrame:
    frame = pd.DataFrame((summary.get("promotion_summary") or {}).get("brands", []))
    if frame.empty:
        return frame
    return frame.rename(
        columns={
            "competitor_name": "Marka",
            "promotion_count": "Kampanya Adedi",
            "basket_discount_count": "Sepette Indirim",
            "rate_discount_count": "Oran Bazli Indirim",
            "installment_count": "Taksit / Finansman",
            "amount_discount_count": "Tutar Bazli Indirim",
            "top_discount_value": "En Yuksek Indirim Degeri",
            "top_discount_unit": "Indirim Birimi",
            "sample_message": "Ornek Mesaj",
        }
    )


def _load_stock_report_frame(summary: dict[str, Any]) -> pd.DataFrame:
    frame = pd.DataFrame(summary.get("stock_summary", []))
    if frame.empty:
        return frame
    return frame.rename(columns={"competitor_name": "Marka", "out_of_stock_count": "Stokta Yok"})


def _build_overview_frame(summary: dict[str, Any]) -> pd.DataFrame:
    overview = summary.get("overview", {})
    price_summary = summary.get("price_summary", {})
    metrics = [
        {"Metrik": "Izlenen Rakip", "Deger": overview.get("competitor_count", 0)},
        {"Metrik": "Toplam Urun", "Deger": overview.get("product_count", 0)},
        {"Metrik": "Haftalik Kampanya", "Deger": overview.get("weekly_promotion_count", 0)},
        {"Metrik": "Stokta Yok", "Deger": overview.get("out_of_stock_count", 0)},
        {"Metrik": "Fiyati Dusan Urun", "Deger": price_summary.get("price_decreased_count", 0)},
        {"Metrik": "Fiyati Artan Urun", "Deger": price_summary.get("price_increased_count", 0)},
        {"Metrik": "Top Discount Marka", "Deger": price_summary.get("top_discount_brand") or "Veri yok"},
    ]
    return pd.DataFrame(metrics)


def build_report_frames(summary: dict[str, Any]) -> dict[str, pd.DataFrame]:
    management_notes = pd.DataFrame({"Yonetici Ozeti": build_management_summary(summary)})
    catalog_diff = pd.DataFrame(flatten_catalog_diff_rows(summary))
    specs = _load_product_specs_report_frame()

    return {
        "overview": _build_overview_frame(summary),
        "management_summary": management_notes,
        "price_changes": _load_price_changes_report_frame(summary),
        "promotions": _load_promotion_report_frame(summary),
        "stock_risk": _load_stock_report_frame(summary),
        "catalog_diff": catalog_diff,
        "product_specs": specs,
    }


def _write_dataframe_sheet(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    if frame.empty:
        pd.DataFrame([{"Durum": "Veri bulunamadi"}]).to_excel(writer, sheet_name=sheet_name, index=False)
        return
    frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    for index, column in enumerate(frame.columns, start=1):
        max_length = max(len(str(column)), *(len(str(value)) for value in frame[column].head(200).tolist()))
        worksheet.column_dimensions[_excel_column_letter(index)].width = min(max_length + 2, 32)


def _style_range(worksheet: Any, start_row: int, end_row: int, start_col: int, end_col: int, fill: Any, bold: bool = False, font_color: str = "000000") -> None:
    assets = _style_assets()
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.fill = fill
            cell.border = assets["THIN_BORDER"]
            cell.font = assets["Font"](bold=bold, color=font_color)
            cell.alignment = assets["Alignment"](horizontal="center", vertical="center", wrap_text=True)


def _auto_fit_columns(worksheet: Any, max_width: int = 28) -> None:
    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells[:250] if cell.value is not None]
        if not values:
            continue
        width = min(max(len(str(value)) for value in values) + 2, max_width)
        worksheet.column_dimensions[_excel_column_letter(column_cells[0].column)].width = width


def _apply_currency_format(worksheet: Any, start_row: int, end_row: int, price_start_col: int, price_end_col: int) -> None:
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=price_start_col, max_col=price_end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = _style_assets()["Alignment"](horizontal="center", vertical="center")


def _write_manual_panel_sheet(workbook: Any, panel_frame: pd.DataFrame, snapshot_columns: list[str]) -> None:
    assets = _style_assets()
    worksheet = workbook.create_sheet("Yemek Odasi")
    worksheet.sheet_properties.tabColor = "70AD47"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet["A1"] = "PANEL REKABET"
    worksheet["A2"] = "Marka, takim ve urun cesidi bazinda otomatik olusan rekabet paneli"
    end_col = 5 + len(snapshot_columns) * 2
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    _style_range(worksheet, 1, 1, 1, end_col, assets["HEADER_FILL"], bold=True, font_color="FFFFFF")
    _style_range(worksheet, 2, 2, 1, end_col, assets["SECTION_FILL"])

    if panel_frame.empty:
        worksheet["A4"] = "Veri bulunamadi."
        return

    static_headers = ["Marka", "Malzeme Kodu", "Takim", "Urun Cesidi", "Urun Adi"]
    for index, header in enumerate(static_headers, start=1):
        worksheet.merge_cells(start_row=3, start_column=index, end_row=4, end_column=index)
        worksheet.cell(3, index, header)
    current_col = len(static_headers) + 1
    for snapshot_label in snapshot_columns:
        worksheet.merge_cells(start_row=3, start_column=current_col, end_row=3, end_column=current_col + 1)
        worksheet.cell(3, current_col, snapshot_label)
        worksheet.cell(4, current_col, "Liste fiyat")
        worksheet.cell(4, current_col + 1, "IND. PRK FIYAT")
        current_col += 2
    _style_range(worksheet, 3, 3, 1, end_col, assets["HEADER_FILL"], bold=True, font_color="FFFFFF")
    _style_range(worksheet, 4, 4, 1, end_col, assets["SUBHEADER_FILL"], bold=True)
    for price_col in range(6, end_col + 1):
        worksheet.cell(4, price_col).fill = assets["PRICE_FILL"]
    worksheet.auto_filter.ref = f"A4:{_excel_column_letter(end_col)}4"
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[3].height = 24
    worksheet.row_dimensions[4].height = 22

    current_row = 5
    display_brands = [brand.upper() for brand in BRAND_ORDER if brand.upper() in panel_frame["Marka"].unique()]
    for brand in display_brands:
        brand_rows = panel_frame.loc[panel_frame["Marka"] == brand].copy()
        if brand_rows.empty:
            continue
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=end_col)
        worksheet.cell(current_row, 1, brand)
        _style_range(worksheet, current_row, current_row, 1, end_col, assets["SECTION_FILL"], bold=True)
        current_row += 1

        for _, row in brand_rows.iterrows():
            row_values = [row["Marka"], row["Malzeme Kodu"], row["Takim"], row["Urun Cesidi"], row["Urun Adi"]]
            for snapshot_label in snapshot_columns:
                row_values.append(row.get(f"{snapshot_label}|Liste fiyat"))
                row_values.append(row.get(f"{snapshot_label}|IND. PRK FIYAT"))
            for col_index, value in enumerate(row_values, start=1):
                cell = worksheet.cell(current_row, col_index, value)
                cell.border = assets["THIN_BORDER"]
                cell.alignment = assets["Alignment"](
                    horizontal="center" if col_index >= 6 else "left",
                    vertical="top",
                    wrap_text=True,
                )
            fill = assets["TEAM_FILL"] if bool(row["Takim Satiri"]) else assets["ITEM_FILL"]
            font = assets["Font"](bold=bool(row["Takim Satiri"]), color="FFFFFF" if bool(row["Takim Satiri"]) else "000000")
            for col_index in range(1, end_col + 1):
                worksheet.cell(current_row, col_index).fill = fill
                worksheet.cell(current_row, col_index).font = font
            worksheet.row_dimensions[current_row].height = 20 if bool(row["Takim Satiri"]) else 36
            current_row += 1
        current_row += 1
    if current_row > 5 and end_col >= 6:
        _apply_currency_format(worksheet, 5, current_row - 1, 6, end_col)
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 38
    _auto_fit_columns(worksheet, max_width=30)


def _write_price_band_sheet(workbook: Any, price_band_frame: pd.DataFrame) -> None:
    assets = _style_assets()
    worksheet = workbook.create_sheet("FIYAT ARALIGI_endeks")
    worksheet.sheet_properties.tabColor = "C9DAF8"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A4"
    worksheet["A1"] = "Fiyat Araligi ve Endeks"
    worksheet["A2"] = "Marka bazli fiyat bandi dagilimi"
    end_col = max(1, len(price_band_frame.columns))
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    _style_range(worksheet, 1, 1, 1, end_col, assets["HEADER_FILL"], bold=True, font_color="FFFFFF")
    _style_range(worksheet, 2, 2, 1, end_col, assets["SECTION_FILL"])

    if price_band_frame.empty:
        worksheet["A4"] = "Veri bulunamadi."
        return

    for index, column in enumerate(price_band_frame.columns, start=1):
        worksheet.cell(4, index, column)
    _style_range(worksheet, 4, 4, 1, end_col, assets["SUBHEADER_FILL"], bold=True)
    for row_offset, (_, row) in enumerate(price_band_frame.iterrows(), start=5):
        for col_index, value in enumerate(row.tolist(), start=1):
            cell = worksheet.cell(row_offset, col_index, value)
            cell.border = assets["THIN_BORDER"]
            cell.alignment = assets["Alignment"](horizontal="center", vertical="center")
        if row_offset % 2 == 1:
            _style_range(worksheet, row_offset, row_offset, 1, end_col, assets["SECTION_FILL"])
    worksheet.auto_filter.ref = f"A4:{_excel_column_letter(end_col)}4"
    _auto_fit_columns(worksheet, max_width=22)


def export_weekly_report_excel(summary: dict[str, Any], frames: dict[str, pd.DataFrame], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    manual_context = build_manual_report_context()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        workbook = writer.book
        default_sheet = workbook.active
        if default_sheet and default_sheet.max_row == 1 and default_sheet["A1"].value is None:
            workbook.remove(default_sheet)

        _write_manual_panel_sheet(workbook, manual_context["panel"], manual_context["snapshot_columns"])
        _write_price_band_sheet(workbook, manual_context["price_bands"])

        for sheet_name, frame in frames.items():
            _write_dataframe_sheet(writer, sheet_name[:31], frame)
    return output_path


def _draw_wrapped_text(text_object: Any, text: str, width: int = 95) -> None:
    line = []
    for word in text.split():
        candidate = " ".join(line + [word])
        if len(candidate) > width and line:
            text_object.textLine(" ".join(line))
            line = [word]
        else:
            line.append(word)
    if line:
        text_object.textLine(" ".join(line))


def export_weekly_report_pdf(summary: dict[str, Any], output_path: Path) -> Path:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_canvas = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4

    report_canvas.setTitle("Haftalik Rakip Analiz Raporu")
    report_canvas.setFont("Helvetica-Bold", 16)
    report_canvas.drawString(40, height - 50, "Haftalik Rakip Analiz Raporu")
    report_canvas.setFont("Helvetica", 10)
    report_canvas.drawString(40, height - 68, f"Uretim Tarihi: {datetime.now(timezone.utc).astimezone().isoformat()}")

    y_position = height - 110
    report_canvas.setFont("Helvetica-Bold", 12)
    report_canvas.drawString(40, y_position, "Yonetici Ozeti")
    y_position -= 18

    text_object = report_canvas.beginText(50, y_position)
    text_object.setFont("Helvetica", 10)
    for note in build_management_summary(summary):
        _draw_wrapped_text(text_object, f"- {note}")
        text_object.textLine("")
    report_canvas.drawText(text_object)

    y_position = text_object.getY() - 12
    if y_position < 180:
        report_canvas.showPage()
        y_position = height - 50

    report_canvas.setFont("Helvetica-Bold", 12)
    report_canvas.drawString(40, y_position, "Kritik Metrikler")
    y_position -= 20
    report_canvas.setFont("Helvetica", 10)

    overview_frame = _build_overview_frame(summary)
    for _, row in overview_frame.iterrows():
        report_canvas.drawString(50, y_position, f"{row['Metrik']}: {row['Deger']}")
        y_position -= 14

    y_position -= 10
    report_canvas.setFont("Helvetica-Bold", 12)
    report_canvas.drawString(40, y_position, "Haftalik Katalog Hareketi")
    y_position -= 18
    report_canvas.setFont("Helvetica", 10)

    diff_rows = flatten_catalog_diff_rows(summary)
    if not diff_rows:
        report_canvas.drawString(50, y_position, "Katalog diff verisi bulunamadi.")
    else:
        for row in diff_rows:
            report_canvas.drawString(
                50,
                y_position,
                (
                    f"{row['Marka']}: +{row['Yeni Urun']} yeni, -{row['Kalkan Urun']} pasif, "
                    f"guncel katalog {row['Guncel Katalog']}"
                ),
            )
            y_position -= 14
            if y_position < 60:
                report_canvas.showPage()
                y_position = height - 50
                report_canvas.setFont("Helvetica", 10)

    report_canvas.save()
    return output_path


def _attachment_content_type(path: Path) -> tuple[str, str]:
    guessed, _ = mimetypes.guess_type(path.name)
    if not guessed:
        return ("application", "octet-stream")
    maintype, subtype = guessed.split("/", 1)
    return maintype, subtype


def _build_email_body(summary: dict[str, Any], metadata: dict[str, Any]) -> str:
    lines = [
        "Haftalik rakip analiz raporu ektedir.",
        "",
        "Yonetici Ozeti:",
    ]
    lines.extend(f"- {item}" for item in build_management_summary(summary))
    lines.extend(
        [
            "",
            f"PDF: {metadata['files']['pdf']['path']}",
            f"Excel: {metadata['files']['excel']['path']}",
        ]
    )
    return "\n".join(lines)


def send_weekly_report_email(summary: dict[str, Any], metadata: dict[str, Any]) -> dict[str, Any]:
    recipients = parse_email_recipients(os.getenv("SMTP_TO"))
    smtp_host = os.getenv("SMTP_HOST")
    smtp_from = os.getenv("SMTP_FROM")

    if not smtp_host or not smtp_from or not recipients:
        return {
            "status": "skipped",
            "reason": "SMTP ayarlari tamamlanmamis",
            "recipients": recipients,
        }

    message = EmailMessage()
    message["Subject"] = "Haftalik Urun Bazli Rakip Analiz Raporu"
    message["From"] = smtp_from
    message["To"] = ", ".join(recipients)
    message.set_content(_build_email_body(summary, metadata))

    for file_info in metadata.get("files", {}).values():
        path = Path(file_info["path"])
        if not path.exists():
            continue
        maintype, subtype = _attachment_content_type(path)
        message.add_attachment(path.read_bytes(), maintype=maintype, subtype=subtype, filename=path.name)

    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = os.getenv("SMTP_USERNAME")
    smtp_password = os.getenv("SMTP_PASSWORD")
    use_ssl = os.getenv("SMTP_USE_SSL", "false").lower() == "true"
    use_tls = os.getenv("SMTP_USE_TLS", "true").lower() == "true"

    if use_ssl:
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as smtp:
            if smtp_username and smtp_password:
                smtp.login(smtp_username, smtp_password)
            smtp.send_message(message)
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as smtp:
            if use_tls:
                smtp.starttls()
            if smtp_username and smtp_password:
                smtp.login(smtp_username, smtp_password)
            smtp.send_message(message)

    return {
        "status": "sent",
        "recipients": recipients,
        "smtp_host": smtp_host,
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }


def generate_weekly_report(send_email: bool = False) -> dict[str, Any]:
    summary = generate_executive_summary()
    frames = build_report_frames(summary)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    excel_path = export_weekly_report_excel(summary, frames, REPORTS_DIR / f"weekly_competitor_report_{timestamp}.xlsx")
    pdf_path = export_weekly_report_pdf(summary, REPORTS_DIR / f"weekly_competitor_report_{timestamp}.pdf")

    metadata: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "management_summary": build_management_summary(summary),
        "files": {
            "excel": {"path": str(excel_path), "name": excel_path.name},
            "pdf": {"path": str(pdf_path), "name": pdf_path.name},
        },
    }

    email_status = None
    if send_email:
        try:
            email_status = send_weekly_report_email(summary, metadata)
        except Exception as exc:
            logger.exception("Weekly report email delivery failed")
            email_status = {"status": "failed", "reason": str(exc)}
    else:
        email_status = {"status": "not_requested"}

    metadata["email_delivery"] = email_status
    dump_json(str(REPORT_METADATA_PATH), metadata)
    logger.info(
        "Weekly reporting completed",
        extra={
            "extra_fields": {
                "excel_report": str(excel_path),
                "pdf_report": str(pdf_path),
                "email_status": email_status.get("status"),
            }
        },
    )
    return metadata
